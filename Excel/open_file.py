from openpyxl import load_workbook  # 파일 불러오기

wb = load_workbook("sample.xlsx")  # sample.xslx 파일에서 wb을 불러옴.
ws = wb.active  # 활성화된 sheet

# cell data 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# cell 개수를 모를 때,
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()