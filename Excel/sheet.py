from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()  # 새로운 Sheet를 기본 이름으로 생성
ws.title = "MySheet"  # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff99ff"  # RGB 형태로 값을 넣어주면 색 변경

ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)  # 2번째 index에 Sheet를 생성

new_ws = wb["NewSheet"]  # Dictionary 형태로 Sheet에 접근
# new_ws....

print(wb.sheetnames)  # 모든 Sheet 이름 확인

# Sheet copy
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

print(wb.sheetnames)

wb.save("sample.xlsx")
